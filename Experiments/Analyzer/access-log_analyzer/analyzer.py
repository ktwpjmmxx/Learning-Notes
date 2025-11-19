import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.dates as mdates
from datetime import datetime
import os

print("=== WiFi分析システム起動 (前月比較機能付き) ===")

# 日本語フォント設定
plt.rcParams['font.family'] = 'MS Gothic'
plt.rcParams['axes.unicode_minus'] = False

class WiFiAnalyzerApp:
    def __init__(self, root):
        print("初期化開始...")
        self.root = root
        self.root.title("WiFiアクセスポイント分析システム")
        self.root.geometry("1400x900")
        self.root.configure(bg='#f0f0f0')
        
        self.current_data = None
        self.previous_data = None
        self.ap_names = []
        self.fig = None
        self.canvas = None
        
        print("UI構築...")
        self.setup_ui()
        print("初期化完了")
        
    def setup_ui(self):
        # ヘッダー
        header_frame = tk.Frame(self.root, bg='#667eea', pady=10)
        header_frame.pack(fill=tk.X)
        
        tk.Label(header_frame, text="WiFi分析システム", font=('Arial', 16, 'bold'), bg='#667eea', fg='white').pack()
        
        # ファイル選択（横並び）
        file_frame = tk.Frame(self.root, bg='#f0f0f0', pady=8)
        file_frame.pack(fill=tk.X)
        
        file_container = tk.Frame(file_frame, bg='#f0f0f0')
        file_container.pack()
        
        # 当月ファイル
        tk.Label(file_container, text="当月:", font=('Arial', 10, 'bold'), bg='#f0f0f0').pack(side=tk.LEFT, padx=3)
        
        # ファイル名表示部分（枠で囲む）
        current_file_frame = tk.Frame(file_container, bg='#ffffff', relief=tk.SOLID, borderwidth=1)
        current_file_frame.pack(side=tk.LEFT, padx=3)
        self.current_file_label = tk.Label(current_file_frame, text="未選択", font=('Arial', 9), bg='#ffffff', fg='#666', width=20, anchor='w', padx=5, pady=2)
        self.current_file_label.pack()
        
        tk.Button(file_container, text="📁", command=self.load_current_file, font=('Arial', 9), bg='#667eea', fg='white', padx=8, pady=3, cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        tk.Label(file_container, text="|", bg='#f0f0f0', fg='#ccc').pack(side=tk.LEFT, padx=8)
        
        # 前月ファイル
        tk.Label(file_container, text="前月:", font=('Arial', 10, 'bold'), bg='#f0f0f0').pack(side=tk.LEFT, padx=3)
        
        # ファイル名表示部分（枠で囲む）
        previous_file_frame = tk.Frame(file_container, bg='#ffffff', relief=tk.SOLID, borderwidth=1)
        previous_file_frame.pack(side=tk.LEFT, padx=3)
        self.previous_file_label = tk.Label(previous_file_frame, text="未選択(任意)", font=('Arial', 9), bg='#ffffff', fg='#999', width=20, anchor='w', padx=5, pady=2)
        self.previous_file_label.pack()
        
        tk.Button(file_container, text="📁", command=self.load_previous_file, font=('Arial', 9), bg='#10b981', fg='white', padx=8, pady=3, cursor='hand2').pack(side=tk.LEFT, padx=3)
        
        # ボタンフレーム
        btn_frame = tk.Frame(self.root, bg='#f0f0f0', pady=5)
        btn_frame.pack()
        
        self.export_btn = tk.Button(btn_frame, text="💾 グラフ保存", command=self.export_chart, font=('Arial', 9), bg='#10b981', fg='white', padx=12, pady=5, cursor='hand2', state=tk.DISABLED)
        self.export_btn.pack(side=tk.LEFT, padx=3)
        
        self.report_btn = tk.Button(btn_frame, text="📄 レポート作成", command=self.export_excel_report, font=('Arial', 9), bg='#f59e0b', fg='white', padx=12, pady=5, cursor='hand2', state=tk.DISABLED)
        self.report_btn.pack(side=tk.LEFT, padx=3)
        
        tk.Label(btn_frame, text="|", bg='#f0f0f0', fg='#ccc').pack(side=tk.LEFT, padx=8)
        
        # オプション
        self.show_ma = tk.BooleanVar(value=False)
        tk.Checkbutton(btn_frame, text="7日平均", variable=self.show_ma, command=self.update_chart, font=('Arial', 9), bg='#f0f0f0').pack(side=tk.LEFT, padx=5)
        
        self.show_grid = tk.BooleanVar(value=True)
        tk.Checkbutton(btn_frame, text="グリッド", variable=self.show_grid, command=self.update_chart, font=('Arial', 9), bg='#f0f0f0').pack(side=tk.LEFT, padx=5)
        
        # メインコンテンツ
        content_frame = tk.Frame(self.root, bg='white')
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)
        
        # 統計カード
        self.stats_frame = tk.Frame(content_frame, bg='white')
        self.stats_frame.pack(fill=tk.X, pady=5)
        
        # グラフエリア
        self.chart_frame = tk.Frame(content_frame, bg='white')
        self.chart_frame.pack(fill=tk.BOTH, expand=True)
        
        # 初期メッセージ
        tk.Label(self.chart_frame, text="当月のExcelファイルを選択してください\n前月データも選択すると前月比較分析が表示されます", font=('Arial', 12), bg='white', fg='#666', justify=tk.CENTER).pack(expand=True)
    
    def load_current_file(self):
        print("当月ファイル選択...")
        file_path = filedialog.askopenfilename(title="当月のExcelファイルを選択", filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        
        if not file_path:
            return
        
        try:
            print(f"ファイル: {file_path}")
            self.current_data = self.load_data_from_file(file_path)
            filename = os.path.basename(file_path)
            if len(filename) > 25:
                filename = filename[:22] + "..."
            self.current_file_label.config(text=filename, fg='#333')
            
            self.update_stats()
            self.update_chart()
            
            self.export_btn.config(state=tk.NORMAL)
            self.report_btn.config(state=tk.NORMAL)
            
            msg = f"{len(self.current_data)}日分のデータを読み込みました"
            if self.previous_data is not None:
                msg += "\n前月比較機能が有効になりました"
            messagebox.showinfo("成功", msg)
            print("当月ファイル読み込み完了")
            
        except Exception as e:
            print(f"エラー: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました:\n{str(e)}")
    
    def load_previous_file(self):
        print("前月ファイル選択...")
        file_path = filedialog.askopenfilename(title="前月のExcelファイルを選択", filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        
        if not file_path:
            return
        
        try:
            print(f"ファイル: {file_path}")
            self.previous_data = self.load_data_from_file(file_path)
            filename = os.path.basename(file_path)
            if len(filename) > 25:
                filename = filename[:22] + "..."
            self.previous_file_label.config(text=filename, fg='#333')
            
            if self.current_data is not None:
                self.update_stats()
                messagebox.showinfo("成功", "前月データを読み込みました\n前月比較機能が有効になりました")
            else:
                messagebox.showinfo("確認", "前月データを読み込みました\n当月データを選択すると比較分析が表示されます")
            
            print("前月ファイル読み込み完了")
            
        except Exception as e:
            print(f"エラー: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました:\n{str(e)}")
    
    def load_data_from_file(self, file_path):
        print("データ読み込み処理開始...")
        df = pd.read_excel(file_path)
        
        if df.empty:
            raise ValueError("ファイルにデータがありません")
        
        columns = df.columns.tolist()
        print(f"列名: {columns}")
        
        date_col = None
        for col in columns:
            if '日時' in str(col) or '日付' in str(col) or 'date' in str(col).lower():
                date_col = col
                break
        
        if date_col is None:
            date_col = columns[0]
        
        ap_columns = [col for col in columns if col != date_col]
        
        if len(ap_columns) < 3:
            raise ValueError(f"アクセスポイントの列が3つ必要です (現在: {len(ap_columns)}列)")
        
        if not self.ap_names:
            self.ap_names = ap_columns[:3]
            # デバッグ: AP名を出力して確認
            print(f"AP名設定: {self.ap_names}")
            for i, name in enumerate(self.ap_names):
                print(f"  AP{i+1}: '{name}' (長さ: {len(str(name))})")
        
        # 日付列を変換
        print(f"日付列の元データ型: {df[date_col].dtype}")
        print(f"日付列の元データ（最初の5行）: {df[date_col].head().tolist()}")
        
        # Excelから読み込んだ日付を変換
        if df[date_col].dtype == 'datetime64[ns]':
            print("日付列は既にdatetime型です")
        elif df[date_col].dtype in ['int64', 'float64']:
            # Excelのシリアル値として変換（1900年1月1日基準）
            print("日付列をExcelシリアル値から変換します")
            df[date_col] = pd.TimedeltaIndex(df[date_col], unit='d') + pd.Timestamp('1899-12-30')
        else:
            # 文字列の場合
            print("日付列を文字列から変換します")
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        
        print(f"変換後の日付型: {df[date_col].dtype}")
        print(f"変換後の日付（最初の5行）: {df[date_col].head().tolist()}")
        
        data = df[[date_col] + ap_columns[:3]].copy()
        data.columns = ['日付', 'AP1', 'AP2', 'AP3']
        
        for col in ['AP1', 'AP2', 'AP3']:
            data[col] = pd.to_numeric(data[col], errors='coerce').fillna(0)
        
        data = data.dropna(subset=['日付'])
        data['曜日'] = data['日付'].dt.dayofweek
        
        data['AP1_MA'] = data['AP1'].rolling(window=7, min_periods=1).mean()
        data['AP2_MA'] = data['AP2'].rolling(window=7, min_periods=1).mean()
        data['AP3_MA'] = data['AP3'].rolling(window=7, min_periods=1).mean()
        
        print(f"データ読み込み完了: {len(data)}行")
        return data
    
    def calculate_comparison(self):
        print("前月比較計算中...")
        if self.previous_data is None:
            return None
        
        comparison = {}
        for col in ['AP1', 'AP2', 'AP3']:
            current_total = self.current_data[col].sum()
            previous_total = self.previous_data[col].sum()
            diff = current_total - previous_total
            diff_percent = (diff / previous_total * 100) if previous_total > 0 else 0
            
            comparison[col] = {
                'current': current_total,
                'previous': previous_total,
                'diff': diff,
                'diff_percent': diff_percent
            }
        
        print("前月比較計算完了")
        return comparison
    
    def update_stats(self):
        print("統計カード更新中...")
        for widget in self.stats_frame.winfo_children():
            widget.destroy()
        
        if self.current_data is None:
            return
        
        totals = [self.current_data['AP1'].sum(), self.current_data['AP2'].sum(), self.current_data['AP3'].sum()]
        averages = [self.current_data['AP1'].mean(), self.current_data['AP2'].mean(), self.current_data['AP3'].mean()]
        
        comparison = self.calculate_comparison()
        
        colors = ['#dbeafe', '#d1fae5', '#fef3c7']
        text_colors = ['#2563eb', '#059669', '#d97706']
        
        for i, (name, total, avg, bg_color, fg_color) in enumerate(zip(self.ap_names, totals, averages, colors, text_colors)):
            card = tk.Frame(self.stats_frame, bg=bg_color, relief=tk.RAISED, borderwidth=1)
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
            
            tk.Label(card, text=name, font=('Arial', 10, 'bold'), bg=bg_color, fg='#333').pack(pady=(8, 2))
            tk.Label(card, text=f"{int(total):,}", font=('Arial', 18, 'bold'), bg=bg_color, fg=fg_color).pack()
            tk.Label(card, text="総アクセス数", font=('Arial', 8), bg=bg_color, fg='#666').pack(pady=(0, 2))
            
            # 前月比を表示
            if comparison:
                col_name = ['AP1', 'AP2', 'AP3'][i]
                comp = comparison[col_name]
                diff = comp['diff']
                diff_percent = comp['diff_percent']
                
                if diff > 0:
                    arrow = "↗"
                    color = '#059669'
                elif diff < 0:
                    arrow = "↘"
                    color = '#dc2626'
                else:
                    arrow = "→"
                    color = '#666'
                
                tk.Label(card, text=f"{arrow} {diff_percent:+.1f}%", font=('Arial', 11, 'bold'), bg=bg_color, fg=color).pack()
                tk.Label(card, text=f"前月: {int(comp['previous']):,}", font=('Arial', 7), bg=bg_color, fg='#999').pack(pady=(0, 2))
            
            tk.Label(card, text=f"平均: {avg:.1f}/日", font=('Arial', 8), bg=bg_color, fg='#666').pack(pady=(0, 8))
        
        print("統計カード更新完了")
    
    def update_chart(self):
        if self.current_data is None:
            return
        
        print("グラフ更新中...")
        for widget in self.chart_frame.winfo_children():
            widget.destroy()
        
        self.fig = Figure(figsize=(12, 6), dpi=100)
        ax = self.fig.add_subplot(111)
        
        dates = self.current_data['日付'].tolist()
        day_of_weeks = self.current_data['曜日'].tolist()
        
        ax.plot(dates, self.current_data['AP1'], 'o-', color='#3b82f6', label=self.ap_names[0], linewidth=2, markersize=4)
        ax.plot(dates, self.current_data['AP2'], 'o-', color='#10b981', label=self.ap_names[1], linewidth=2, markersize=4)
        ax.plot(dates, self.current_data['AP3'], 'o-', color='#f59e0b', label=self.ap_names[2], linewidth=2, markersize=4)
        
        if self.show_ma.get():
            ax.plot(dates, self.current_data['AP1_MA'], '--', color='#3b82f6', alpha=0.5, linewidth=1.5, label=f'{self.ap_names[0]} (7日平均)')
            ax.plot(dates, self.current_data['AP2_MA'], '--', color='#10b981', alpha=0.5, linewidth=1.5, label=f'{self.ap_names[1]} (7日平均)')
            ax.plot(dates, self.current_data['AP3_MA'], '--', color='#f59e0b', alpha=0.5, linewidth=1.5, label=f'{self.ap_names[2]} (7日平均)')
        
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
        ax.xaxis.set_major_locator(mdates.DayLocator(interval=1))
        
        for i, (label, day) in enumerate(zip(ax.get_xticklabels(), day_of_weeks)):
            if day == 6:
                label.set_color('#2563eb')
            elif day == 0:
                label.set_color('#dc2626')
        
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        if self.show_grid.get():
            ax.grid(True, alpha=0.3, linestyle='--')
        
        ax.set_xlabel('日付', fontsize=11, fontweight='bold')
        ax.set_ylabel('アクセス数', fontsize=11, fontweight='bold')
        
        title = f'アクセス数推移グラフ ({len(self.current_data)}日分)'
        if self.previous_data is not None:
            title += ' - 前月比較データあり'
        ax.set_title(title, fontsize=13, fontweight='bold', pad=15)
        
        ax.legend(loc='upper left', framealpha=0.9, fontsize=9)
        self.fig.tight_layout(pad=2.0)
        
        self.canvas = FigureCanvasTkAgg(self.fig, self.chart_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        self.setup_hover_tooltip(ax)
        print("グラフ更新完了")
    
    def setup_hover_tooltip(self, ax):
        self.annot = ax.annotate("", xy=(0,0), xytext=(20,20), textcoords="offset points", bbox=dict(boxstyle="round", fc="white", ec="gray", alpha=0.95), arrowprops=dict(arrowstyle="->"))
        self.annot.set_visible(False)
        
        def on_hover(event):
            if event.inaxes != ax:
                if self.annot.get_visible():
                    self.annot.set_visible(False)
                    self.canvas.draw_idle()
                return
            
            if event.xdata is None or event.ydata is None:
                return
            
            dates = mdates.date2num(self.current_data['日付'].tolist())
            x_mouse = event.xdata
            idx = (abs(dates - x_mouse)).argmin()
            
            date_str = self.current_data.iloc[idx]['日付'].strftime('%Y/%m/%d')
            ap1_val = int(self.current_data.iloc[idx]['AP1'])
            ap2_val = int(self.current_data.iloc[idx]['AP2'])
            ap3_val = int(self.current_data.iloc[idx]['AP3'])
            total = ap1_val + ap2_val + ap3_val
            
            text = f"{date_str}\n{self.ap_names[0]}: {ap1_val:,}\n{self.ap_names[1]}: {ap2_val:,}\n{self.ap_names[2]}: {ap3_val:,}\n合計: {total:,}"
            
            self.annot.xy = (dates[idx], event.ydata)
            self.annot.set_text(text)
            self.annot.set_visible(True)
            self.canvas.draw_idle()
        
        self.canvas.mpl_connect("motion_notify_event", on_hover)
    
    def export_chart(self):
        if self.fig is None:
            return
        
        file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG files", "*.png"), ("PDF files", "*.pdf")])
        
        if file_path:
            try:
                self.fig.savefig(file_path, dpi=300, bbox_inches='tight')
                messagebox.showinfo("成功", f"グラフを保存しました:\n{file_path}")
            except Exception as e:
                messagebox.showerror("エラー", f"保存に失敗しました:\n{str(e)}")
    
    def export_excel_report(self):
        if self.current_data is None:
            return
        
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        
        if file_path:
            temp_files = []
            try:
                print("Excelレポート作成中...")
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    self.create_combined_summary_sheet(writer)
                    temp_file = self.create_chart_sheet(writer)
                    if temp_file:
                        temp_files.append(temp_file)
                
                for tf in temp_files:
                    try:
                        import time
                        time.sleep(0.2)
                        os.unlink(tf)
                        print(f"一時ファイル削除成功: {tf}")
                    except Exception as e:
                        print(f"一時ファイル削除失敗（問題ありません）: {e}")
                
                messagebox.showinfo("成功", f"月次レポートを保存しました:\n{file_path}")
                print("Excelレポート作成完了")
            except Exception as e:
                print(f"Excel作成エラー: {e}")
                import traceback
                traceback.print_exc()
                for tf in temp_files:
                    try:
                        os.unlink(tf)
                    except:
                        pass
                messagebox.showerror("エラー", f"保存に失敗しました:\n{str(e)}")
    
    def create_combined_summary_sheet(self, writer):
        print("統合サマリーシート作成中...")
        summary_data = []
        
        summary_data.append(['WiFi アクセスポイント 月次レポート'])
        summary_data.append([''])
        
        start_date = self.current_data['日付'].min().strftime('%Y/%m/%d')
        end_date = self.current_data['日付'].max().strftime('%Y/%m/%d')
        summary_data.append(['【基本情報】'])
        summary_data.append(['分析期間', f'{start_date} - {end_date}'])
        summary_data.append(['日数', f'{len(self.current_data)}日間'])
        summary_data.append([''])
        
        summary_data.append(['【当月統計】'])
        summary_data.append(['項目', self.ap_names[0], self.ap_names[1], self.ap_names[2], '合計'])
        
        totals = [int(self.current_data['AP1'].sum()), int(self.current_data['AP2'].sum()), int(self.current_data['AP3'].sum())]
        total_all = sum(totals)
        summary_data.append(['総アクセス数', totals[0], totals[1], totals[2], total_all])
        
        avgs = [self.current_data['AP1'].mean(), self.current_data['AP2'].mean(), self.current_data['AP3'].mean()]
        avg_all = sum(avgs)
        summary_data.append(['平均/日', f'{avgs[0]:.1f}', f'{avgs[1]:.1f}', f'{avgs[2]:.1f}', f'{avg_all:.1f}'])
        
        max_vals = []
        for col in ['AP1', 'AP2', 'AP3']:
            max_idx = self.current_data[col].idxmax()
            max_date = self.current_data.loc[max_idx, '日付'].strftime('%m/%d')
            max_val = int(self.current_data.loc[max_idx, col])
            max_vals.append(f'{max_val} ({max_date})')
        summary_data.append(['最大値(日付)', max_vals[0], max_vals[1], max_vals[2], ''])
        
        min_vals = []
        for col in ['AP1', 'AP2', 'AP3']:
            min_idx = self.current_data[col].idxmin()
            min_date = self.current_data.loc[min_idx, '日付'].strftime('%m/%d')
            min_val = int(self.current_data.loc[min_idx, col])
            min_vals.append(f'{min_val} ({min_date})')
        summary_data.append(['最小値(日付)', min_vals[0], min_vals[1], min_vals[2], ''])
        
        stds = [self.current_data['AP1'].std(), self.current_data['AP2'].std(), self.current_data['AP3'].std()]
        summary_data.append(['標準偏差', f'{stds[0]:.1f}', f'{stds[1]:.1f}', f'{stds[2]:.1f}', ''])
        
        summary_data.append([''])
        
        if self.previous_data is not None:
            comparison = self.calculate_comparison()
            
            summary_data.append(['【前月比較】'])
            summary_data.append(['項目', self.ap_names[0], self.ap_names[1], self.ap_names[2]])
            
            current_totals = [int(comparison['AP1']['current']), int(comparison['AP2']['current']), int(comparison['AP3']['current'])]
            summary_data.append(['当月総数', current_totals[0], current_totals[1], current_totals[2]])
            
            previous_totals = [int(comparison['AP1']['previous']), int(comparison['AP2']['previous']), int(comparison['AP3']['previous'])]
            summary_data.append(['前月総数', previous_totals[0], previous_totals[1], previous_totals[2]])
            
            diffs = [int(comparison['AP1']['diff']), int(comparison['AP2']['diff']), int(comparison['AP3']['diff'])]
            summary_data.append(['差分', diffs[0], diffs[1], diffs[2]])
            
            diff_percents = [f"{comparison['AP1']['diff_percent']:+.1f}%", f"{comparison['AP2']['diff_percent']:+.1f}%", f"{comparison['AP3']['diff_percent']:+.1f}%"]
            summary_data.append(['増減率', diff_percents[0], diff_percents[1], diff_percents[2]])
            
            total_current = sum(current_totals)
            total_previous = sum(previous_totals)
            total_diff_percent = ((total_current - total_previous) / total_previous * 100) if total_previous > 0 else 0
            
            if total_diff_percent > 10:
                status = '大幅増加 - 容量確認推奨'
            elif total_diff_percent > 0:
                status = '増加傾向'
            elif total_diff_percent < -10:
                status = '大幅減少 - 原因調査推奨'
            else:
                status = '安定推移'
            
            summary_data.append([''])
            summary_data.append(['総評', status])
            summary_data.append(['全体増減率', f'{total_diff_percent:+.1f}%'])
            summary_data.append([''])
        
        summary_data.append(['【曜日別平均アクセス数】'])
        summary_data.append(['曜日', '平均アクセス数', '最多曜日'])
        weekday_names = ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日', '土曜日', '日曜日']
        weekday_data = []
        for day in range(7):
            day_data = self.current_data[self.current_data['曜日'] == day]
            if len(day_data) > 0:
                day_avg = day_data[['AP1', 'AP2', 'AP3']].sum(axis=1).mean()
                weekday_data.append((weekday_names[day], day_avg))
        
        max_day_name = max(weekday_data, key=lambda x: x[1])[0] if weekday_data else ''
        
        for name, avg in weekday_data:
            mark = '★' if name == max_day_name else ''
            summary_data.append([name, f'{avg:.1f}', mark])
        
        summary_data.append([''])
        
        summary_data.append(['【アクセスポイント別構成比】'])
        summary_data.append(['AP名', 'アクセス数', '構成比'])
        for i, name in enumerate(self.ap_names):
            ratio = (totals[i] / total_all * 100) if total_all > 0 else 0
            summary_data.append([name, int(totals[i]), f'{ratio:.1f}%'])
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='統計レポート', index=False, header=False)
        
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        worksheet = writer.sheets['統計レポート']
        worksheet.column_dimensions['A'].width = 22
        worksheet.column_dimensions['B'].width = 18
        worksheet.column_dimensions['C'].width = 18
        worksheet.column_dimensions['D'].width = 18
        worksheet.column_dimensions['E'].width = 18
        
        # スタイル定義
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        section_font = Font(bold=True, size=10)
        border_style = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # タイトル行のスタイル（1行目）
        for col in range(1, 6):
            cell = worksheet.cell(1, col)
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        worksheet.merge_cells('A1:E1')
        
        # 特定行の処理（セル結合と色付け）
        merge_rows = [3, 7, 15, 25, 35]
        header_rows = [8, 16, 26, 36]
        
        # セル結合と色付け
        for row_num in merge_rows:
            if row_num <= worksheet.max_row:
                worksheet.merge_cells(f'A{row_num}:E{row_num}')
                cell = worksheet.cell(row_num, 1)
                cell.fill = section_fill
                cell.font = section_font
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border_style
        
        # 青背景・白文字
        for row_num in header_rows:
            if row_num <= worksheet.max_row:
                for col in range(1, 6):
                    cell = worksheet.cell(row_num, col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_style
        
        # 全セルに枠線とスタイルを適用
        for row_idx in range(2, worksheet.max_row + 1):
            if row_idx == 1 or row_idx in merge_rows or row_idx in header_rows:
                continue
                
            for col_idx in range(1, 6):
                cell = worksheet.cell(row_idx, col_idx)
                cell.border = border_style
                
                if cell.value is not None and cell.value != '':
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        if isinstance(cell.value, (int, float)):
                            if isinstance(cell.value, float) and cell.value.is_integer():
                                cell.value = int(cell.value)
                            cell.number_format = '0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
        
        print("統合サマリーシート作成完了")
    
    def create_chart_sheet(self, writer):
        print("グラフシート作成中...")
        if self.previous_data is not None:
            fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(18, 5))
        else:
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
        
        # グラフ1: 円グラフ
        totals = [self.current_data['AP1'].sum(), self.current_data['AP2'].sum(), self.current_data['AP3'].sum()]
        colors = ['#1F77B4', '#FF7F0E', '#2CA02C']
        
        wedges, texts, autotexts = ax1.pie(
            totals, 
            colors=colors, 
            autopct='%1.1f%%', 
            startangle=90,
            textprops={'fontsize': 10, 'weight': 'bold'},
            explode=(0.01, 0.01, 0.01)
        )
        
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_weight('bold')
        
        legend = ax1.legend(
            wedges, 
            [f'{name}  ' for name in self.ap_names],
            title="アクセスポイント",
            loc="center left",
            bbox_to_anchor=(1.05, 0.5),
            fontsize=10,
            frameon=False,
            labelspacing=0.8
        )
        legend.get_title().set_fontweight('bold')
        legend.get_title().set_fontsize(10)
        for text in legend.get_texts():
            text.set_horizontalalignment('center')  # 中央揃えに変更
        
        start_date = self.current_data['日付'].min().strftime('%m/%d')
        end_date = self.current_data['日付'].max().strftime('%m/%d')
        ax1.set_title(f'AP別構成比 ({start_date}～{end_date})', fontsize=12, fontweight='bold', pad=12)
        
        # グラフ2: 曜日別平均
        weekday_names = ['月', '火', '水', '木', '金', '土', '日']
        weekday_avgs = []
        for day in range(7):
            day_data = self.current_data[self.current_data['曜日'] == day]
            if len(day_data) > 0:
                day_avg = day_data[['AP1', 'AP2', 'AP3']].sum(axis=1).mean()
                weekday_avgs.append(day_avg)
            else:
                weekday_avgs.append(0)
        
        weekday_colors = ['#5f6368', '#5f6368', '#5f6368', '#5f6368', '#5f6368', '#4285f4', '#ea4335']
        bars = ax2.bar(weekday_names, weekday_avgs, color=weekday_colors, alpha=0.85)
        ax2.set_title('曜日別平均アクセス数', fontsize=12, fontweight='bold', pad=12)
        ax2.set_ylabel('アクセス数', fontsize=10)
        ax2.grid(axis='y', alpha=0.3)
        
        for bar, val in zip(bars, weekday_avgs):
            height = bar.get_height()
            ax2.text(bar.get_x() + bar.get_width()/2., height + max(weekday_avgs)*0.03,
                    f'{val:.0f}',
                    ha='center', va='bottom', fontsize=9, fontweight='bold')
        
        ax2.set_ylim(0, max(weekday_avgs) * 1.15)
        
        # グラフ3: 前月比較
        if self.previous_data is not None:
            comparison = self.calculate_comparison()
            x = [0, 1, 2]  # 明示的に位置を指定
            width = 0.35
            
            current_vals = [comparison['AP1']['current'], comparison['AP2']['current'], comparison['AP3']['current']]
            previous_vals = [comparison['AP1']['previous'], comparison['AP2']['previous'], comparison['AP3']['previous']]
            
            # 棒の位置を計算
            x1 = [i - width/2 for i in x]
            x2 = [i + width/2 for i in x]
            
            bars1 = ax3.bar(x1, current_vals, width, label='当月', color='#3b82f6', alpha=0.8)
            bars2 = ax3.bar(x2, previous_vals, width, label='前月', color='#94a3b8', alpha=0.8)
            
            ax3.set_xlabel('アクセスポイント', fontsize=10)
            ax3.set_ylabel('総アクセス数', fontsize=10)
            ax3.set_title('前月比較', fontsize=12, fontweight='bold', pad=12)
            
            # X軸の目盛りとラベルを明示的に設定
            ax3.set_xticks(x)
            ax3.set_xticklabels(self.ap_names, fontsize=10, ha='center')
            
            ax3.legend(fontsize=9)
            ax3.grid(axis='y', alpha=0.3)
            
            max_val = max(max(current_vals), max(previous_vals))
            
            for bar, val in zip(bars1, current_vals):
                height = bar.get_height()
                ax3.text(bar.get_x() + bar.get_width()/2., height + max_val*0.02,
                        f'{int(val):,}',
                        ha='center', va='bottom', fontsize=8, fontweight='bold')
            
            for bar, val in zip(bars2, previous_vals):
                height = bar.get_height()
                ax3.text(bar.get_x() + bar.get_width()/2., height + max_val*0.02,
                        f'{int(val):,}',
                        ha='center', va='bottom', fontsize=8, fontweight='bold')
            
            ax3.set_ylim(0, max_val * 1.12)
        
        plt.tight_layout(pad=1.5)
        
        import tempfile
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        temp_file.close()
        fig.savefig(temp_file.name, dpi=150, bbox_inches='tight')
        plt.close(fig)
        
        empty_df = pd.DataFrame([['グラフは下に表示されています']])
        empty_df.to_excel(writer, sheet_name='グラフ', index=False, header=False)
        
        from openpyxl.drawing.image import Image as OpenpyxlImage
        worksheet = writer.sheets['グラフ']
        img = OpenpyxlImage(temp_file.name)
        worksheet.add_image(img, 'A3')
        
        print("グラフシート作成完了")
        return temp_file.name

def main():
    print("メイン関数開始")
    try:
        root = tk.Tk()
        print("Tkinter初期化完了")
        app = WiFiAnalyzerApp(root)
        print("アプリ初期化完了、メインループ開始")
        root.mainloop()
        print("=== 正常終了 ===")
    except Exception as e:
        print(f"=== エラー発生 ===")
        print(f"エラー: {e}")
        import traceback
        traceback.print_exc()
        input("Enterで終了...")

if __name__ == "__main__":
    main()